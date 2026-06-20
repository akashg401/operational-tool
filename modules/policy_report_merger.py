import io
import pandas as pd
from openpyxl import load_workbook



REQUIRED_COLUMNS = [
    "Policy Number",
    "Traveller Name",
    "Policy Path",
    "Policy Amount",
    "Passport",
]


OPTIONAL_COLUMNS = [
    
    "Hub Name",
]


KEEP_COLUMNS = (
    REQUIRED_COLUMNS +
    OPTIONAL_COLUMNS
)


from openpyxl import load_workbook
import pandas as pd


def read_policy_report(uploaded_file):

    wb = load_workbook(
        uploaded_file,
        data_only=True
    )

    ws = wb["Policy Details"]

    records = []

    for row in range(2, ws.max_row + 1):

        policy_path_cell = ws.cell(row, 6)

        policy_url = ""

        if policy_path_cell.hyperlink:
            policy_url = (
                policy_path_cell.hyperlink.target
            )

        hub_name = ""

        if ws.max_column >= 8:
            hub_name = ws.cell(
                row,
                8
            ).value

        records.append({

            "Policy Number":
                ws.cell(row, 1).value,

            "Issue Date":
                ws.cell(row, 2).value,

            "Asego Partner":
                ws.cell(row, 3).value,

            "Traveller Name":
                ws.cell(row, 4).value,

            "Policy Amount":
                ws.cell(row, 5).value,

            "Policy Path":
                "🔗 Click Here",

            "Policy Path URL":
                policy_url,

            "Passport":
                ws.cell(row, 7).value,

            "Hub Name":
                hub_name

        })

    return pd.DataFrame(records)


def merge_policy_reports(files):

    all_frames = []

    total_rows = 0

    for file in files:

        df = read_policy_report(file)

        total_rows += len(df)

        

        all_frames.append(df)

    merged_df = pd.concat(
        all_frames,
        ignore_index=True
    )

    before = len(merged_df)

    merged_df = merged_df.drop_duplicates(
        subset=["Policy Number"],
        keep="first"
    )

    duplicates_removed = (
        before - len(merged_df)
    )

    summary = {
        "files_uploaded": len(files),
        "rows_read": total_rows,
        "duplicates_removed": duplicates_removed,
        "final_records": len(merged_df),
    }

    return merged_df, summary


from openpyxl import Workbook


def export_merged_report(df):

    output = io.BytesIO()

    wb = Workbook()

    ws = wb.active

    ws.title = "Policy Details"

    export_df = df.drop(
        columns=["Policy Path URL"],
        errors="ignore"
    )

    for col_idx, column in enumerate(
        export_df.columns,
        start=1
    ):
        ws.cell(
            row=1,
            column=col_idx,
            value=column
        )

    for row_idx, row in enumerate(
        export_df.itertuples(index=False),
        start=2
    ):

        for col_idx, value in enumerate(
            row,
            start=1
        ):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

    policy_path_col = list(
        export_df.columns
    ).index("Policy Path") + 1

    for row_idx, url in enumerate(
        df["Policy Path URL"],
        start=2
    ):

        if url:

            ws.cell(
                row=row_idx,
                column=policy_path_col
            ).hyperlink = url

    wb.save(output)

    output.seek(0)

    return output.getvalue()
